##reading existing xlsx file

from openpyxl import load_workbook
#for handling file
import os


path = './source'



diff = {
    0:'No Difference in Windchill and QAD',
    1:'No Values Present in Windchill',
    2:'Part is not Effective Windchill',
    3:'Difference in Effectivity Start',
    4:'Difference in Effectivity End',
    5:'Difference in Structure',
    6:'Differnece in Quantity',
    7:'No Values Present in QAD',
    8:'Difference in Parent Number',
    9:'Difference in Child Number'
    }

#Secondary workbook loader
def checkXlsx(book):
    wbs = load_workbook(filename="./source/" + book)
    wss = wbs.active
    last=9
    output_individual_Sheet = set()
    
    while(1):
        if (wss.cell(row=last, column=1).value == None):
            break
        else:
            last+= 1
    if(last == 9):
        output_individual_Sheet.add(0)
        print(output_individual_Sheet)
        return(output_individual_Sheet) 
    else:
        for i in range(9, last):

            if(wss.cell(row=i, column=1).value=='No Values Present in Windchill'):
                output_individual_Sheet.add(1)

            elif(wss.cell(row=i, column=1).value=='Part is not Effective Windchill'):
                output_individual_Sheet.add(2)
            
            elif(wss.cell(row=i, column=8).value=='No Values Present in QAD'):
                output_individual_Sheet.add(7)

            else:
                if(wss.cell(row=i, column=2).value!=wss.cell(row=i, column=8).value):
                    output_individual_Sheet.add(8)
                if(wss.cell(row=i, column=3).value!=wss.cell(row=i, column=9).value):
                    output_individual_Sheet.add(9)
                if(wss.cell(row=i, column=4).value!=wss.cell(row=i, column=10).value):
                    output_individual_Sheet.add(3)
                if(wss.cell(row=i, column=5).value!=wss.cell(row=i, column=11).value):
                    output_individual_Sheet.add(4)
                if(wss.cell(row=i, column=6).value!=wss.cell(row=i, column=12).value):
                    output_individual_Sheet.add(5)
                if(wss.cell(row=i, column=7).value!=wss.cell(row=i, column=13).value):
                    output_individual_Sheet.add(6)
        print(output_individual_Sheet)
        return output_individual_Sheet

    

def openFile(keyword):
#save files as list
    files = os.listdir(path)
    
    for file in files:
        if file.startswith(keyword):
            print (file)
            x = checkXlsx(file)
            print(x)
            return x


#Primary workbook loader
wb = load_workbook(filename="Bom_Compare.xlsx")

ws = wb.active
part_name = ws['C']

# print the content

for x in range(1,len(part_name)):
    print(part_name[x].value)
    se = openFile(part_name[x].value)
    print('-------------------------------------')
    print(se)
    st = str()
    count = 0
    for r in se:
        if count:
            st += ' & '
        st += diff[r]
        count += 1
    print('in cell = '+st)
    ws['E'+str(x+1)] = st
    print ('E'+str(x+1))


wb.save('Result.xlsx')